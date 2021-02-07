from copy import Error
from inspect import cleandoc
from openpyxl import Workbook, load_workbook
from openpyxl.utils import FORMULAE
import tkinter as tk
from tkinter import ttk
from tkinter.filedialog import askopenfilename

EXECL_PATH = ""

OUT_WORKBOOK_SUFFIX = "_out"
DEBUG_LABEL = None


def update_debug_label(txt, color="red"):
    DEBUG_LABEL.config(fg=color, text=txt)


def get_out_filepath():
    prefix, delimiter, postfix = EXECL_PATH.rpartition(".")
    return prefix + "_out" + delimiter + postfix


def is_client_row(sheet, cell):
    client_name = [r for r in sheet.iter_rows(
        min_row=cell.row, max_row=cell.row)][0][0].value
    return client_name != None


def calculate_spreadsheet():
    if(EXECL_PATH == ""):
        update_debug_label("You didn't select a table!")
        return

    table_workbook = load_workbook(
        filename=EXECL_PATH, data_only=True)
    # loop thorough each sheet and multiply each column by its price
    for table_sheet in table_workbook.worksheets:
        # we save the max_column so when we create the sum cell the max_columns number don't grow
        table_sheet_max_columns = table_sheet.max_column
        for col in table_sheet.iter_cols(min_row=1, min_col=1, max_row=table_sheet.max_row, max_col=table_sheet_max_columns):
            # if the column don't have a price, skip the column
            if not (type(col[0].value) == int or type(col[0].value) == float):
                continue
            col_price = col[0].value

            # iterate through all the cell of current column and multiply by the price_id
            # we start at the third row, because the first two are the id and the product name
            # if this row don't have a client name we skip it except the last row, whice is a sum row of each column
            for cell in col[2:]:
                cellvalue_first_column_of_current_row = [r for r in table_sheet.iter_rows(
                    min_row=cell.row, max_row=cell.row)][0][0].value
                if not is_client_row(table_sheet, cell):
                    continue

                cell.value = cell.value if cell.value else 0
                cell.value = 0 if cell.value == ' ' else cell.value
                try:
                    cell.value = float(cell.value)*float(col_price)
                except:
                    print(
                        f"========\ncan't convert the cell[{cell.row},{cell.column_letter}]: '{cell.value}' to float. (sheet title: {table_sheet.title})\n=========")
                    raise ValueError()
                # create or upated the sum cell of each row
                sum_cell = table_sheet.cell(
                    row=cell.row, column=table_sheet_max_columns+1)
                if sum_cell.value != None:
                    sum_cell.value = sum_cell.value + cell.value
                else:
                    sum_cell.value = 0 + cell.value

    # collect the sum data of each client for each sheet
    clients_sum_data = {}
    for table_sheet in table_workbook.worksheets:
        for row in table_sheet.iter_rows(min_row=3, min_col=1, max_col=table_sheet.max_column, max_row=table_sheet.max_row):
            if(not is_client_row(table_sheet, row[0])):
                continue
            if row[0].value in clients_sum_data:
                clients_sum_data[row[0].value][table_sheet.title] = row[-1].value
            else:
                clients_sum_data[row[0].value] = {
                    table_sheet.title: row[-1].value}

    for client in clients_sum_data:
        try:
            clients_sum_data[client]["summery"] = sum(
                clients_sum_data[client].values())
        except:
            print(f"{client}: {clients_sum_data[client]}")

    # create a summery sheet and dump here the data
    summery_sheet = table_workbook.create_sheet("summery")
    summery_sheet.cell(row=1, column=1).value = "Clients"

    dict_sheettitle_to_colnum = {}
    col_counter = 2
    for sheet in table_workbook.worksheets:
        summery_sheet.cell(row=1, column=col_counter).value = sheet.title
        dict_sheettitle_to_colnum[sheet.title] = col_counter
        col_counter += 1

    summery_row = 2
    for client in sorted(clients_sum_data):
        summery_sheet.cell(row=summery_row, column=1).value = client
        for sheet in clients_sum_data[client]:
            summery_sheet.cell(
                row=summery_row, column=dict_sheettitle_to_colnum[sheet]).value = clients_sum_data[client][sheet]
        summery_row += 1

    # Save all sheets to new workbook
    table_workbook.save(filename=get_out_filepath())
    update_debug_label("Done!", "green")


def create_frame(root, label_text):
    load_frame = tk.Frame(master=root, relief=tk.GROOVE, borderwidth=5)
    tk.Label(master=load_frame,
             text=label_text).grid(row=1, column=1, padx=5, pady=5)
    load_label = tk.Label(text="", master=load_frame, width=30)
    load_label.grid(row=2, column=1, padx=5, pady=5)

    # This is where we lauch the file manager bar.
    def OpenFile(label):
        global EXECL_PATH
        filename = askopenfilename(initialdir="./",
                                   filetypes=(("Execl File", "*.xlsx"),
                                              ("All Files", "*.*")),
                                   title="Choose a file."
                                   )
        label_text = filename
        if(len(filename) > 30):
            label_text = "..." + filename[-30:]

        label.config(text=label_text)
        EXECL_PATH = filename

    tk.Button(master=load_frame, text="Browse",
              command=lambda: OpenFile(load_label)).grid(row=2, column=2, padx=5, pady=5)

    return load_frame


def tk_window():
    root = tk.Tk()
    root.title("SpreadSheeter")

    global DEBUG_LABEL
    if(DEBUG_LABEL == None):
        DEBUG_LABEL = tk.Label(
            text="Click Finish to create the SpreadSheeted table")

    DEBUG_LABEL.grid(row=4, column=1, padx=5, pady=5)

    table_frame = create_frame(
        root, "Please select the execl file to SpreadSheet: ")
    table_frame.grid(row=1, column=1, padx=20, pady=10)

    tk.Button(text="Finish",
              command=calculate_spreadsheet).grid(row=3, column=1, padx=20, pady=10)

    root.mainloop()


def main():
    tk_window()


if __name__ == "__main__":
    main()
